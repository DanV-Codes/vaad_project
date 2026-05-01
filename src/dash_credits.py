import streamlit as st
import pandas as pd
import os
import openpyxl

# --- פונקציית שמירה לאקסל ---
def save_collection_to_excel(df_to_save, master_file):
    try:
        wb = openpyxl.load_workbook(master_file)
        ws = wb['גביה 2026']
        for index, row in df_to_save.iterrows():
            excel_row = index + 2
            # השמירה מתבצעת לפי שמות העמודות, כך שסדר התצוגה בדשבורד לא משפיע על המיקום באקסל
            ws.cell(row=excel_row, column=2).value = row.get('דירה', '')
            if 'שם משפחה' in row:
                ws.cell(row=excel_row, column=3).value = row['שם משפחה']
            if 'קופה קטנה' in row:
                ws.cell(row=excel_row, column=4).value = row['קופה קטנה']
            
            months_heb = ['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני', 'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר']
            for i, month in enumerate(months_heb):
                if month in row:
                    ws.cell(row=excel_row, column=5 + i).value = row[month]
        
        wb.save(master_file)
        return True
    except Exception as e:
        st.error(f"שגיאה בשמירה לאקסל: {e}")
        return False

# --- פונקציה לטעינת חריגים ---
def get_exceptions_data(month_name, exceptions_dir):
    MONTHS_MAP = {
        'ינואר': 1, 'פברואר': 2, 'מרץ': 3, 'אפריל': 4,
        'מאי': 5, 'יוני': 6, 'יולי': 7, 'אוגוסט': 8,
        'ספטמבר': 9, 'אוקטובר': 10, 'נובמבר': 11, 'דצמבר': 12
    }
    month_num = MONTHS_MAP.get(month_name, 1)
    file_path = os.path.join(exceptions_dir, f'exceptions_month_{month_num}.csv')
    try:
        if os.path.exists(file_path):
            return pd.read_csv(file_path, encoding='utf-8-sig'), file_path
    except Exception as e:
        st.error(f"שגיאה בקריאת קובץ החריגים: {e}")
    return None, file_path

# --- הפונקציה הראשית ---
def render_credits(selected_month_name, MONTHS_HEB, MASTER_FILE, COMMITTEE_APTS, DEBT_LIMIT=350, PETTY_LIMIT=500):
    st.header(f"💰 ניהול גביית דיירים")
    
    # טעינת 60 הדירות בלבד כדי להימנע משורות סיכום (כמו "סה"כ")
    if st.session_state.master_df is not None:
        df = st.session_state.master_df.head(60)
    else:
        return

    # 1. סטטוס גבייה עליון
    col1, col2 = st.columns(2)
    with col1:
        if selected_month_name in df.columns:
            unpaid_count = df[~df['דירה'].isin(COMMITTEE_APTS) & ((df[selected_month_name] < DEBT_LIMIT) | (df[selected_month_name].isna()))].shape[0]
            st.metric(label=f"דירות עם חוב ב{selected_month_name}", value=unpaid_count)
    
    with col2:
        if 'קופה קטנה' in df.columns:
            unpaid_petty = df[((df['קופה קטנה'] < PETTY_LIMIT) | (df['קופה קטנה'].isna()))].shape[0]
            st.metric(label="דירות עם חוב לקופה קטנה", value=unpaid_petty)

    # 2. עריכת טבלה
    st.subheader('עריכת נתוני גבייה')
    
    # הגדרת סדר העמודות: קופה קטנה משמאל, חודשים באמצע, דירה הכי ימנית
    display_cols = ['קופה קטנה'] + MONTHS_HEB + ['שם משפחה', 'דירה']
    
    # סינון העמודות שקיימות בנתונים בפועל
    actual_cols = [col for col in display_cols if col in df.columns]
    
    edited_table = st.data_editor(
        df[actual_cols],
        use_container_width=True,
        height=400,
        hide_index=True,
        key="collection_ed"
    )
    
    if st.button('💾 שמור שינויים באקסל'):
        if save_collection_to_excel(edited_table, MASTER_FILE):
            st.success("הנתונים נשמרו!")
            st.session_state.master_df = edited_table
            st.rerun()

    st.divider()

    # 3. הודעות ווטסאפ וחריגים
    st.subheader("📱 הודעות מוכנות וטיפול בחריגים")
    col_whatsapp, col_exceptions = st.columns(2)
    
    with col_whatsapp:
        # חובות ועד שוטף
        st.markdown("💬 **חובות ועד שוטף**")
        current_month_idx = MONTHS_HEB.index(selected_month_name)
        selected_whatsapp_months = st.multiselect(
            'בחר חודשים לבדיקה:', MONTHS_HEB, default=MONTHS_HEB[:current_month_idx + 1]
        )
        
        if selected_whatsapp_months:
            msg_lines = []
            df_paying = df[~df['דירה'].isin(COMMITTEE_APTS)].copy()
            for _, row in df_paying.iterrows():
                try:
                    apt_val = row['דירה']
                    if pd.isna(apt_val) or str(apt_val).strip() == '': continue
                    apt = int(apt_val)
                    debts_for_apt = []
                    for m in selected_whatsapp_months:
                        val = row[m] if pd.notna(row.get(m)) else 0
                        if val < DEBT_LIMIT:
                            debts_for_apt.append(f"{m} (חוב {int(DEBT_LIMIT - val)})")
                    if debts_for_apt:
                        msg_lines.append(f"דירה {apt} - {', '.join(debts_for_apt)}")
                except: continue
            
            if msg_lines:
                st.code("\n".join(msg_lines), language='text')
            else:
                st.success("אין חובות שוטפים!")

        st.write("")
        # חובות קופה קטנה
        st.markdown("💬 **חובות קופה קטנה (500 ש״ח)**")
        if 'קופה קטנה' in df.columns:
            petty_msg = []
            for _, row in df.iterrows():
                try:
                    apt_val = row['דירה']
                    if pd.isna(apt_val) or str(apt_val).strip() == '': continue
                    apt = int(apt_val)
                    val = row['קופה קטנה'] if pd.notna(row.get('קופה קטנה')) else 0
                    if val < PETTY_LIMIT:
                        petty_msg.append(f"דירה {apt} - קופה קטנה (חוב {int(PETTY_LIMIT - val)})")
                except: continue
            
            if petty_msg:
                st.code("\n".join(petty_msg), language='text')
            else:
                st.success("גביית קופה קטנה הושלמה מכולם!")

    # ניהול חריגים
    with col_exceptions:
        st.markdown(f"⚠️ **טיפול בחריגים לחודש {selected_month_name}**")
        exceptions_dir = os.path.join('data', 'exceptions')
        full_ex_df, ex_file_path = get_exceptions_data(selected_month_name, exceptions_dir)
        
        if full_ex_df is not None and not full_ex_df.empty:
            st.info("סמן ב-**V** שורות שטיפלת בהן, ולחץ על כפתור המחיקה.")
            ex_display = ['Date', 'Amount', 'Category', 'Apartment_Number', 'Entity_Name']
            existing_ex = [c for c in ex_display if c in full_ex_df.columns]
            display_df = full_ex_df[existing_ex].copy()
            display_df.insert(0, 'טופל (למחיקה)', False)
            
            edited_ex = st.data_editor(display_df, hide_index=True, use_container_width=True, key=f"ex_editor_{selected_month_name}")
            
            if st.button("🗑️ מחק חריגים שסומנו כ'טופל'"):
                rows_to_keep = edited_ex[edited_ex['טופל (למחיקה)'] == False].index
                updated_full_df = full_ex_df.loc[rows_to_keep]
                updated_full_df.to_csv(ex_file_path, index=False, encoding='utf-8-sig')
                st.success("החריגים עודכנו בהצלחה!")
                st.rerun()
        else:
            st.success("אין חריגים ממתינים לטיפול בחודש זה. 🎉")