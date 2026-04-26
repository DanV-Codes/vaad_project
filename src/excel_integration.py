import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import os
import tkinter as tk
from tkinter import scrolledtext

# =======================================================
# חלון 1: סיכום גבייה (עיצוב מקורי + יישור לימין)
# =======================================================

def show_summary_popup(month_num, updated_count, already_filled_apts, newly_updated_apts, unidentified_df, amount_exceptions_df):
    """
    חלון קופץ עם סיכום הרצת הגבייה - שומר על העיצוב המקורי עם יישור לימין.
    """
    root = tk.Tk()
    root.title(f"דוח סיכום גבייה - חודש {month_num}")
    root.geometry("850x650")
    root.configure(bg="#f4f4f4")

    def close_app():
        root.quit()
        root.destroy()

    title_lbl = tk.Label(root, text=f"📊 סיכום הרצת נתונים באקסל - חודש {month_num}", font=("Arial", 18, "bold"), bg="#f4f4f4")
    title_lbl.pack(pady=15)

    stats_frame = tk.Frame(root, bg="#f4f4f4")
    stats_frame.pack(fill=tk.X, padx=20)

    success_text = f"✅ עודכנו בהצלחה: {updated_count} דירות"
    if newly_updated_apts:
        success_text += f" ({sorted(newly_updated_apts)})"
    tk.Label(stats_frame, text=success_text, font=("Arial", 14), fg="green", bg="#f4f4f4").pack(anchor="e", pady=2)

    if already_filled_apts:
        tk.Label(stats_frame, text=f"ℹ️ דולגו (כבר שילמו באקסל): {len(already_filled_apts)} דירות ({sorted(already_filled_apts)})", font=("Arial", 14), fg="#b8860b", bg="#f4f4f4").pack(anchor="e", pady=2)

    text_area = scrolledtext.ScrolledText(root, wrap=tk.NONE, font=("Courier", 11), bg="#2b2b2b", fg="#a9b7c6")
    text_area.pack(pady=15, padx=20, fill=tk.BOTH, expand=True)
    
    # הגדרת יישור לימין
    text_area.tag_configure("rtl", justify='right')

    if not unidentified_df.empty:
        text_area.insert(tk.END, f"❌ רשימה שחורה (דירה לא זוהתה): {len(unidentified_df)} שורות\n", "rtl")
        text_area.insert(tk.END, "="*70 + "\n", "rtl")
        display_cols = [c for c in ['Date', 'Amount', 'Original_Desc', 'Entity_Name'] if c in unidentified_df.columns]
        text_area.insert(tk.END, unidentified_df[display_cols].to_string(index=False) + "\n\n\n", "rtl")

    if not amount_exceptions_df.empty:
        text_area.insert(tk.END, f"⚠️ חריגי תשלום (סכום לא תואם לחוקים): {len(amount_exceptions_df)} שורות\n", "rtl")
        text_area.insert(tk.END, "="*70 + "\n", "rtl")
        display_cols = [c for c in ['Apartment_Number', 'Amount', 'Date', 'Entity_Name'] if c in amount_exceptions_df.columns]
        text_area.insert(tk.END, amount_exceptions_df[display_cols].to_string(index=False) + "\n", "rtl")

    if unidentified_df.empty and amount_exceptions_df.empty:
        text_area.insert(tk.END, "\n🎉 אין חריגים או שגיאות! הכל עבר חלק.\n", "rtl")

    text_area.config(state=tk.DISABLED)

    close_btn = tk.Button(root, text="סגור דוח והמשך", command=close_app, font=("Arial", 14, "bold"), bg="#4CAF50", fg="white", width=20, pady=10)
    close_btn.pack(pady=15)

    root.mainloop()

# =======================================================
# חלון 2: סיכום הוצאות (3 רשימות + יישור לימין)
# =======================================================

def show_expenses_popup(month_num, updated_log, duplicates_log, not_found_log):
    """
    חלון סיכום הוצאות הכולל שלוש קטגוריות ויישור טקסט לימין.
    """
    root = tk.Tk()
    root.title(f"דוח סיכום הוצאות - חודש {month_num}")
    root.geometry("850x650")
    root.configure(bg="#f4f4f4")

    def close_app():
        root.quit()
        root.destroy()

    tk.Label(root, text=f"💸 סיכום הרצת הוצאות - חודש {month_num}", font=("Arial", 18, "bold"), bg="#f4f4f4").pack(pady=15)
    
    text_area = scrolledtext.ScrolledText(root, wrap=tk.NONE, font=("Courier", 11), bg="#2b2b2b", fg="#a9b7c6")
    text_area.pack(pady=15, padx=20, fill=tk.BOTH, expand=True)
    
    # הגדרת יישור לימין
    text_area.tag_configure("rtl", justify='right')
    
    output = ""
    
    # 1. רשימת הצלחות
    if updated_log:
        output += f"✅ הוצאות שעודכנו בהצלחה ({len(updated_log)}):\n" + "-"*60 + "\n"
        output += "\n".join(updated_log) + "\n\n\n"
    
    # 2. רשימת כפילויות (עם הניסוח שביקשת)
    if duplicates_log:
        output += f"⚠️ כפילויות שנמצאו - לא בוצע עדכון ({len(duplicates_log)}):\n" + "-"*60 + "\n"
        output += "\n".join(duplicates_log) + "\n\n\n"
        
    # 3. רשימת קטגוריות שלא קיימות באקסל
    if not_found_log:
        output += f"❌ קטגוריות שלא נמצאו בטאב ההוצאות ({len(not_found_log)}):\n" + "-"*60 + "\n"
        output += "\n".join(not_found_log)

    text_area.insert(tk.END, output if output else "לא נמצאו תנועות לעדכון.", "rtl")
    text_area.config(state=tk.DISABLED)

    tk.Button(root, text="סגור דוח והמשך", command=close_app, font=("Arial", 14, "bold"), bg="#4CAF50", fg="white", width=20, pady=10).pack(pady=15)
    root.mainloop()

# =======================================================
# פונקציות העדכון
# =======================================================

def update_master_excel(df, master_path):
    """עדכון טאב גבייה 2026"""
    if df.empty: return

    numeric_apt = pd.to_numeric(df['Apartment_Number'], errors='coerce')
    unidentified_df = df[numeric_apt.isna()].copy()
    identified_df = df[~numeric_apt.isna()].copy()
    identified_df['Apt_Numeric'] = numeric_apt[~numeric_apt.isna()]

    try:
        wb = openpyxl.load_workbook(master_path)
        ws = wb["גביה 2026"]
    except Exception as e:
        print(f"❌ שגיאה בטעינת האקסל: {e}"); return

    valid_dates = pd.to_datetime(df['Date'], dayfirst=True, errors='coerce').dropna()
    month_num = valid_dates.iloc[0].month
    target_col = month_num + 4 

    updated_count, newly_updated_apts, already_filled_apts, exceptions_idx = 0, [], [], []

    for idx, row in identified_df.iterrows():
        apt_num, amount = int(row['Apt_Numeric']), row['Amount']
        target_row = apt_num + 1
        current_val = ws.cell(row=target_row, column=target_col).value
        curr_str = str(current_val).strip() if current_val is not None else ""
        if curr_str.endswith(".0"): curr_str = curr_str[:-2]

        if apt_num in [12, 40]:
            if amount in [320, 350] and curr_str in ["", "30"]:
                ws.cell(row=target_row, column=target_col).value = 350
                updated_count += 1; newly_updated_apts.append(apt_num)
            elif curr_str == "350":
                already_filled_apts.append(apt_num)
            else:
                exceptions_idx.append(idx)
        else:
            if amount == 350:
                if curr_str == "":
                    ws.cell(row=target_row, column=target_col).value = 350
                    updated_count += 1; newly_updated_apts.append(apt_num)
                elif curr_str == "350":
                    already_filled_apts.append(apt_num)
                else:
                    exceptions_idx.append(idx)
            else:
                exceptions_idx.append(idx)

    # שמירת חריגים
    amount_exceptions_df = identified_df.loc[exceptions_idx].copy()
    if not unidentified_df.empty or not amount_exceptions_df.empty:
        project_root = os.path.dirname(master_path)
        exc_dir = os.path.join(project_root, "data", "exceptions")
        os.makedirs(exc_dir, exist_ok=True)
        pd.concat([unidentified_df, amount_exceptions_df]).to_csv(os.path.join(exc_dir, f"exceptions_month_{month_num}.csv"), index=False, encoding='utf-8-sig')

    # גיבוי גולמי
    backup_name = f"פירוט_גביה_{month_num}"
    if backup_name in wb.sheetnames: del wb[backup_name]
    ws_backup = wb.create_sheet(title=backup_name)
    for r in dataframe_to_rows(df, index=False, header=True): ws_backup.append(r)

    wb.save(master_path)
    show_summary_popup(month_num, updated_count, already_filled_apts, newly_updated_apts, unidentified_df, amount_exceptions_df)


def update_expenses_in_excel(df_debits, master_path):
    """עדכון טאב הוצאות 2026 עם ניהול כפילויות מפורט"""
    if df_debits.empty: return

    try:
        wb = openpyxl.load_workbook(master_path)
        ws = wb["הוצאות 2026"]
    except Exception as e:
        print(f"❌ שגיאה בטעינת האקסל: {e}"); return

    valid_dates = pd.to_datetime(df_debits['Date'], dayfirst=True, errors='coerce').dropna()
    month_num = valid_dates.iloc[0].month
    target_col = month_num + 1 

    updated_log, duplicates_log, not_found_log = [], [], []
    
    cat_map = {str(ws.cell(row=r, column=1).value).strip(): r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value}

    for _, row in df_debits.iterrows():
        cat = str(row['Category']).strip()
        amt = abs(row['Amount'])

        if cat in cat_map:
            row_idx = cat_map[cat]
            current_val = ws.cell(row=row_idx, column=target_col).value
            
            if current_val in [None, "", 0, "0", "0.0"]:
                ws.cell(row=row_idx, column=target_col).value = amt
                updated_log.append(f"✅ {cat}: עודכן סכום של {amt} ₪")
            else:
                # הניסוח שביקשת לכפילויות
                duplicates_log.append(f"⚠️ {cat}: כפילות נמצאה, רשום סכום {current_val} ולכן לא מעדכן את הסכום {amt}")
        else:
            not_found_log.append(f"❌ {cat}: סכום {amt} ₪ (לא נמצא באקסל)")

    # גיבוי גולמי
    backup_name = f"פירוט_הוצאות_{month_num}"
    if backup_name in wb.sheetnames: del wb[backup_name]
    ws_backup = wb.create_sheet(title=backup_name)
    for r in dataframe_to_rows(df_debits, index=False, header=True): ws_backup.append(r)

    wb.save(master_path)
    show_expenses_popup(month_num, updated_log, duplicates_log, not_found_log)