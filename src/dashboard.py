import streamlit as st
import pandas as pd
import os
import openpyxl
import datetime

# הגדרות עמוד
st.set_page_config(page_title='ניהול ועד - האגמית 7', layout='wide', page_icon='🏢')

# --- קבועים ---
DEBT_LIMIT = 350
PETTY_CASH_LIMIT = 500  # היעד לקופה קטנה
COMMITTEE_APTS = [19, 33, 26, 38]
TOTAL_PAYING_APTS = 60 - len(COMMITTEE_APTS)
MONTHS_HEB = ['ינואר', 'פברואר', 'מרץ', 'אפריל', 'מאי', 'יוני', 'יולי', 'אוגוסט', 'ספטמבר', 'אוקטובר', 'נובמבר', 'דצמבר']
MASTER_FILE = 'האגמית7_כספים_2026.xlsx'
EXCEPTIONS_DIR = os.path.join('data', 'exceptions')
COLLECTION_SHEET = 'גביה 2026'
EXPENSES_SHEET = 'הוצאות 2026'

st.title('📊 מערכת ניהול ועד בית - האגמית 7')

# --- פונקציות טעינה ושמירה ---

def load_collection_data():
    try:
        df = pd.read_excel(MASTER_FILE, sheet_name=COLLECTION_SHEET, usecols='B,E:Q', nrows=60)
        df.columns = ['דירה'] + MONTHS_HEB + ['קופה קטנה']
        
        for col in MONTHS_HEB + ['קופה קטנה']:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        df['דירה'] = df['דירה'].astype(int)
        return df
    except Exception as e:
        st.error(f"שגיאה בקריאת נתוני גבייה: {e}")
        return None

def load_expenses_data():
    try:
        df = pd.read_excel(MASTER_FILE, sheet_name=EXPENSES_SHEET, usecols='A:M')
        df.columns = ['ספק'] + MONTHS_HEB
        for month in MONTHS_HEB:
            df[month] = pd.to_numeric(df[month], errors='coerce').fillna(0)
        return df
    except Exception as e:
        st.error(f"שגיאה בקריאת נתוני הוצאות: {e}")
        return None

def save_collection_to_excel(updated_rows):
    try:
        wb = openpyxl.load_workbook(MASTER_FILE)
        ws = wb[COLLECTION_SHEET]
        
        col_mapping = {name: 5 + i for i, name in enumerate(MONTHS_HEB)}
        col_mapping['קופה קטנה'] = 17 
        
        for _, row in updated_rows.iterrows():
            apt_num = int(row['דירה'])
            target_row = None
            for r in range(2, 65):
                if ws.cell(row=r, column=2).value == apt_num:
                    target_row = r
                    break
            if target_row:
                for col_name, col_idx in col_mapping.items():
                    ws.cell(row=target_row, column=col_idx).value = int(row[col_name])
        wb.save(MASTER_FILE)
        return True
    except Exception as e:
        st.error(f"שגיאה בשמירה לאקסל (גבייה): {e}")
        return False

def save_expenses_to_excel(updated_df):
    try:
        wb = openpyxl.load_workbook(MASTER_FILE)
        ws = wb[EXPENSES_SHEET]
        for index, row in updated_df.iterrows():
            supplier_name = row['ספק']
            target_row = None
            for r in range(2, 50): 
                if ws.cell(row=r, column=1).value == supplier_name:
                    target_row = r
                    break
            if target_row:
                for i, month in enumerate(MONTHS_HEB):
                    ws.cell(row=target_row, column=2 + i).value = row[month]
        wb.save(MASTER_FILE)
        return True
    except Exception as e:
        st.error(f"שגיאה בשמירה לאקסל (הוצאות): {e}")
        return False

# אתחול Session State
if 'master_df' not in st.session_state:
    st.session_state.master_df = load_collection_data()
if 'expenses_df' not in st.session_state:
    st.session_state.expenses_df = load_expenses_data()

# --- תפריט צדדי ---
st.sidebar.header('⚙️ הגדרות')
current_month_index = datetime.datetime.now().month - 1
selected_month_name = st.sidebar.selectbox('בחר חודש לעבודה:', MONTHS_HEB, index=current_month_index)
selected_month_idx = MONTHS_HEB.index(selected_month_name)

# יצירת הטאבים
tab_collection, tab_expenses = st.tabs(['💰 ניהול גבייה', '💸 הוצאות וספקים'])

# --- טאב גבייה ---
with tab_collection:
    if st.session_state.master_df is not None:
        # לווטסאפ: מסננים את חברי הוועד החוצה
        paying_df = st.session_state.master_df[~st.session_state.master_df['דירה'].isin(COMMITTEE_APTS)].copy()
        
        # 1. ווטסאפ
        st.subheader('📱 הודעה מוכנה לווטסאפ')
        default_months_for_report = MONTHS_HEB[:current_month_index + 1]
        report_months = st.multiselect('כלול בהודעה:', MONTHS_HEB, default=default_months_for_report, key='ws_months')
        
        debt_lines = []
        for _, row in paying_df.iterrows():
            apt_debts = []
            for month in report_months:
                if row[month] < DEBT_LIMIT:
                    apt_debts.append(f'{month} (חוב {int(DEBT_LIMIT - row[month])})')
            if apt_debts:
                debt_lines.append(f"דירה {int(row['דירה'])} - {', '.join(apt_debts)}")
        
        whatsapp_message = "היי, להלן רשימת החובות:\n\n" + "\n".join(debt_lines)
        st.code(whatsapp_message, language='text')

        # 2. חריגים
        st.markdown('---')
        st.subheader(f'⚠️ חריגים - {selected_month_name}')
        exc_path = os.path.join(EXCEPTIONS_DIR, f'exceptions_month_{selected_month_idx+1}.csv')
        if os.path.exists(exc_path):
            exc_df = pd.read_csv(exc_path)
            edited_exc = st.data_editor(exc_df, num_rows="dynamic", use_container_width=True, key="exc_ed", hide_index=True)
            if st.button('💾 שמור שינויים בחריגים'):
                edited_exc.to_csv(exc_path, index=False)
                st.success("קובץ החריגים עודכן.")
                st.rerun()
        else:
            st.info(f"לא נמצא קובץ חריגים לחודש {selected_month_name}.")

        # 3. טבלת גבייה
        st.markdown('---')
        st.subheader('📅 עריכת נתוני גבייה')
        
        stats = {'דירה': 'אחוז גבייה %'}
        expected_regular_total = TOTAL_PAYING_APTS * DEBT_LIMIT # 56 * 350
        expected_petty_cash_total = 60 * PETTY_CASH_LIMIT       # 60 * 500
        
        # חישוב אחוזים לחודשים רגילים (ללא דירות ועד)
        for col in MONTHS_HEB:
            collected_sum = pd.to_numeric(paying_df[col], errors='coerce').fillna(0).sum()
            stats[col] = round((collected_sum / expected_regular_total) * 100, 1)
            
        # חישוב אחוזים לקופה קטנה (כל 60 הדירות)
        collected_petty = pd.to_numeric(st.session_state.master_df['קופה קטנה'], errors='coerce').fillna(0).sum()
        stats['קופה קטנה'] = round((collected_petty / expected_petty_cash_total) * 100, 1)
        
        # כדי שנוכל לערוך את הקופה הקטנה של דירות הוועד, מציגים בטבלה את כ-ל הדירות (master_df)
        view_df = st.session_state.master_df.copy()
        view_df['דירה'] = view_df['דירה'].astype(str)
        final_display_df = pd.concat([view_df, pd.DataFrame([stats])], ignore_index=True)
        
        display_cols = ['קופה קטנה'] + MONTHS_HEB[::-1] + ['דירה']

        edited_table = st.data_editor(final_display_df[display_cols], use_container_width=True, height=400, hide_index=True, key="master_ed")

        if st.button('💾 שמור שינויים באקסל הגבייה'):
            rows_to_save = edited_table.iloc[:-1].copy()
            rows_to_save['דירה'] = rows_to_save['דירה'].astype(int)
            if save_collection_to_excel(rows_to_save):
                st.success("נתוני הגבייה והקופה הקטנה נשמרו!")
                st.session_state.master_df = load_collection_data()
                st.rerun()

# --- טאב הוצאות ---
with tab_expenses:
    if st.session_state.expenses_df is not None:
        st.subheader(f'💸 הוצאות ספקים - שנת 2026')
        
        current_month_total = st.session_state.expenses_df[selected_month_name].sum()
        st.metric(label=f"סה״כ הוצאות ל{selected_month_name}", value=f"₪{current_month_total:,.2f}")
        
        exp_display_cols = ['ספק'] + MONTHS_HEB
        
        edited_expenses = st.data_editor(
            st.session_state.expenses_df[exp_display_cols],
            use_container_width=True,
            height=500,
            hide_index=True,
            key="expenses_ed"
        )
        
        if st.button('💾 שמור שינויים בהוצאות'):
            if save_expenses_to_excel(edited_expenses):
                st.success("נתוני ההוצאות נשמרו בהצלחה!")
                st.session_state.expenses_df = load_expenses_data()
                st.rerun()
        
        st.markdown('---')
        st.subheader('📈 מגמת הוצאות חודשית')
        monthly_totals = st.session_state.expenses_df[MONTHS_HEB].sum()
        st.line_chart(monthly_totals)